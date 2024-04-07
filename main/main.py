import os
import sys
import openpyxl
import time
from sort.radix_sort import RadixSort
from sort.comb_sort import CombSort

sys.path.insert(0, os.path.abspath(os.curdir))
from arquivos.manipular_arquivos import Arquivo


arquivo = Arquivo()
radix_sort = RadixSort()
comb_sort = CombSort()


def calcular_tempo_radix(lista: list) -> float:
    inicio = time.time()
    radix_sort.radixSort(lista)
    fim = time.time()
    return fim - inicio


def calcular_tempo_comb(lista: list) -> float:
    inicio = time.time()
    comb_sort.comb_sort(lista)
    fim = time.time()
    return fim - inicio


wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "Radix Sort"

for i in range(1, 16):
    ws1.cell(row=1, column=i, value=f"Lista {i * 10000}")


for i in range(40):
    print(f"iniciou o {i+1}º")
    lista_radix_sort: list = [
        arquivo.carregar_arquivo(f"./arquivos/positivos/lista{i}Mil.txt")
        for i in range(10, 151, 10)
    ]
    for index, lista in enumerate(lista_radix_sort, start=1):
        tempo = calcular_tempo_radix(lista)
        print(tempo)
        ws1.cell(row=i + 2, column=index, value=tempo)

print("Finalizado")

ws2 = wb.create_sheet(title="Comb Sort")

for i in range(1, 16):
    ws2.cell(row=1, column=i, value=f"Lista {i * 10000}")


for i in range(40):
    print(f"iniciou o {i+1}º")
    lista_comb_sort: list = [
        arquivo.carregar_arquivo(f"./arquivos/aleatorios/lista{i}Mil.txt")
        for i in range(10, 151, 10)
    ]

    for index, lista in enumerate(lista_comb_sort, start=1):
        tempo = calcular_tempo_comb(lista)
        print(tempo)
        ws2.cell(row=i + 2, column=index, value=tempo)

print("Finalizado")

wb.save("resultados.xlsx")
