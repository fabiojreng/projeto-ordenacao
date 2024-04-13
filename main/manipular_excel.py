import openpyxl
from calcular_tempo import CalcularTempo
import os, sys

sys.path.insert(0, os.path.abspath(os.curdir))
from arquivos.manipular_arquivos import Arquivo


class ManipularExcel:
    def __init__(self) -> None:
        self.wb = openpyxl.Workbook()
        self.__gerar_excel()

    def __gerar_excel(self) -> None:
        self.planilha1 = self.wb.active
        self.planilha1.title = "Radix Sort"
        self.planilha2 = self.wb.create_sheet(title="Comb Sort")

        for planilha in [self.planilha1, self.planilha2]:
            for i in range(1, 16):
                planilha.cell(row=1, column=i, value=f"Lista {i * 10000}")

    @staticmethod
    def __gerar_valores_planilha(algoritmo, planilha, caminho_arquivo: str):
        for i in range(40):
            print(f"iniciou a {i+1}ª execução")
            lista_sort = Arquivo.carregar_listas(caminho_arquivo)
            for index, lista in enumerate(lista_sort, start=1):
                tempo = CalcularTempo.calcular_tempo(algoritmo, lista)
                print(tempo)
                planilha.cell(row=i + 2, column=index, value=tempo)

    def execute_radix(self, algoritmo) -> None:
        self.__gerar_valores_planilha(
            algoritmo, self.planilha2, "./arquivos/positivos/"
        )

    def execute_comb(self, algoritmo) -> None:
        self.__gerar_valores_planilha(
            algoritmo, self.planilha2, "./arquivos/aleatorios/"
        )

    def salvar(self, nome_arquivo) -> None:
        self.wb.save(nome_arquivo)
